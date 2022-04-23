import openpyxl
import time

begin = time.time()
fileA = openpyxl.load_workbook("附件A 订购方案数据结果.xlsx")
fileB = openpyxl.load_workbook("附件B 转运方案数据结果.xlsx")

for sheet in range(0, 2):
    sheetA = fileA.worksheets[sheet]
    sheetB = fileB.worksheets[sheet]

    order = [3, 6, 2, 8, 4, 1, 7, 5]

    for col in range(2, 26):    # 每周遍历
        count = [0, 0, 0, 0, 0, 0, 0, 0]
        for row in range(7, 409):   # 遍历每家供应商
            if sheetA.cell(row, col).value is None or sheetA.cell(row, col).value == 0:
                continue
            if sheetA.cell(row, col).value <= 6000:
                for index in range(0, 8):   # 按优先级顺序遍历转运商
                    if sheetA.cell(row, col).value + count[index] <= 6000:
                        rowb = 8 * (col - 2) + 1 + order[index]
                        sheetB.cell(row, rowb).value = sheetA.cell(row, col).value
                        count[index] += sheetA.cell(row, col).value
                        break
                    else:
                        index += 1
            else:
                cache = sheetA.cell(row, col).value
                while cache > 6000:
                    for index in range(0, 8):
                        if count[index] == 0:
                            rowb = 8 * (col - 2) + 1 + order[index]
                            sheetB.cell(row, rowb).value = 6000
                            count[index] = 6000
                            break
                        else:
                            index += 1
                    cache -= 6000
                for index in range(0, 8):
                    if cache + count[index] <= 6000:
                        rowb = 8 * (col - 2) + 1 + order[index]
                        sheetB.cell(row, rowb).value = cache
                        count[index] += cache
                        break
                    else:
                        index += 1

fileB.save("附件B 转运方案数据结果NEW.xlsx")
print("done")
end = time.time()
print(end-begin)
