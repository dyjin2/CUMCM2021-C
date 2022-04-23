import openpyxl

filename = "附件1 近5年402家供应商的相关数据.xlsx"
file = openpyxl.load_workbook(filename)
dinghuo = file.worksheets[0]
gonghuo = file.worksheets[1]

wb = openpyxl.Workbook()
wb.create_sheet('processing', 0)
sheet = wb['processing']
# sheet['A1'] = 'hello'
initial = ['供应商', '类型', '订单数', '完成订单数', '完成率', '总供应量', '订单分', '完成分', '总量分', '总分']
sheet.append(initial)

# for row in dinghuo.iter_rows(2, 402, 3, 242):
#     dingdan = 0
#     for cell in row:
#         # print(cell.coordinate, cell.value)
#         if cell.value > 0:
#             dingdan += 1
#     print(cell.coordinate, dingdan)

for row in range(2, 404):
    sheet.cell(row, 1).value = dinghuo.cell(row, 1).value  # S001
    sheet.cell(row, 2).value = dinghuo.cell(row, 2).value  # A
    total = 0
    dingdan = 0
    deal = 0  # abs((ding-gong)/ding) < 0.1
    # count1 = 0
    # count2 = 0
    for col in range(3, 243):
        ding = dinghuo.cell(row, col).value
        if ding == 0:
            continue

        dingdan += 1
        # count1 += 1
        gong = gonghuo.cell(row, col).value
        total += gong
        if abs((ding - gong) / ding) < 0.1:
            deal += 1
        # count2 += 1
    sheet.cell(row, 3).value = dingdan
    sheet.cell(row, 4).value = deal
    sheet.cell(row, 5).value = deal / dingdan
    sheet.cell(row, 6).value = total
    # print(dinghuo.cell(row, 1).value, dinghuo.cell(row, 2).value, dingdan, deal, deal/dingdan, total)

dingdanmax = 0
for row in sheet.iter_rows(min_row=2, min_col=3, max_col=3, max_row=404, values_only=True):
    for value in row:
        if value == None:
            continue
        if dingdanmax < value:
            dingdanmax = value

dealmax = 0
for row in sheet.iter_rows(min_row=2, min_col=4, max_col=4, max_row=404, values_only=True):
    for value in row:
        if value == None:
            continue
        if dealmax < value:
            dealmax = value

rate = 0
for row in sheet.iter_rows(min_row=2, min_col=5, max_col=5, max_row=404, values_only=True):
    for value in row:
        if value == None:
            continue
        if rate < value:
            rate = value

totalmax = 0
for row in sheet.iter_rows(min_row=2, min_col=6, max_col=6, max_row=404, values_only=True):
    for value in row:
        if value == None:
            continue
        if totalmax < value:
            totalmax = value

# print(dingdanmax, rate, totalmax)

sort = {}
for row in range(2, 404):
    sheet.cell(row, 7).value = sheet.cell(row, 3).value / dingdanmax
    sheet.cell(row, 8).value = sheet.cell(row, 5).value / rate
    sheet.cell(row, 9).value = sheet.cell(row, 6).value / totalmax
    sheet.cell(row, 10).value = sheet.cell(row, 7).value + sheet.cell(row, 8).value + sheet.cell(row, 9).value
    sort[sheet.cell(row, 1).value] = sheet.cell(row, 10).value

wb.save('问题一.xlsx')
print(sorted(sort.items(), key=lambda kv: (kv[1], kv[0]), reverse=True))
# sorted(sort.items(), key=lambda kv: (kv[1], kv[0]), reverse=True)
# num = 0
# for item in sort:
#     print(item)
#     num += 1
#     if num == 50:
#         break
